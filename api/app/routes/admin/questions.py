# api/app/routes/admin/questions.py
#
# Rotas administrativas para o banco de questões compartilhado.
# Restrito a super_admin via JWT.
# ─────────────────────────────────────────────────────────────────────────────

import hashlib
import io
import json
import zipfile
from datetime import datetime
from uuid import uuid4

import boto3
from botocore.exceptions import ClientError
from flask import Blueprint, current_app, jsonify, request
from flask_jwt_extended import get_jwt, get_jwt_identity, jwt_required
from sqlalchemy import func

from app.extensions import db, redis_client
from app.models.question import (
    Alternative,
    DifficultyLevel,
    Question,
    QuestionSourceType,
    QuestionTag,
    QuestionType,
    ReviewStatus,
)
from app.models.tenant import Tenant
from app.models.user import UserRole

admin_questions_bp = Blueprint("admin_questions", __name__)

# TTL dos jobs no Redis: 24 horas
_JOB_TTL = 86_400


# ── Helper de autorização ─────────────────────────────────────────────────────


def _require_super_admin():
    claims = get_jwt()
    if claims.get("role") != UserRole.SUPER_ADMIN.value:
        return jsonify({"error": "forbidden", "message": "Acesso restrito."}), 403
    return None


# ── Mapeamentos de normalização ───────────────────────────────────────────────

_DIFFICULTY_MAP = {
    "easy": DifficultyLevel.EASY,
    "medium": DifficultyLevel.MEDIUM,
    "hard": DifficultyLevel.HARD,
    "facil": DifficultyLevel.EASY,
    "fácil": DifficultyLevel.EASY,
    "medio": DifficultyLevel.MEDIUM,
    "médio": DifficultyLevel.MEDIUM,
    "dificil": DifficultyLevel.HARD,
    "difícil": DifficultyLevel.HARD,
}

_QTYPE_MAP = {
    "interpretacao": QuestionType.INTERPRETACAO,
    "interpretação": QuestionType.INTERPRETACAO,
    "aplicacao": QuestionType.APLICACAO,
    "aplicação": QuestionType.APLICACAO,
    "raciocinio": QuestionType.RACIOCINIO,
    "raciocínio": QuestionType.RACIOCINIO,
    "memorizacao": QuestionType.MEMORIZACAO,
    "memorização": QuestionType.MEMORIZACAO,
    "definicao": QuestionType.MEMORIZACAO,
    "definição": QuestionType.MEMORIZACAO,
}

_XLSX_COL_MAP = {
    "disciplina": "discipline",
    "enunciado": "statement",
    "imagem": "image_file",
    "fonte": "context",
    "alternativa a": "alt_a",
    "alternativa b": "alt_b",
    "alternativa c": "alt_c",
    "alternativa d": "alt_d",
    "alternativa e": "alt_e",
    "gabarito": "correct_alternative_key",
    "topico": "topic",
    "tópico": "topic",
    "subtopico": "subtopic",
    "subtópico": "subtopic",
    "dificuldade": "difficulty",
    "banca": "exam_board",
    "ano": "exam_year",
    "concurso": "exam_name",
    "dica": "tip",
    "justificativa": "correct_justification",
    "justificativa gabarito": "correct_justification",
    "justificativa_gabarito": "correct_justification",
    "justificativa a": "just_a",
    "justificativa b": "just_b",
    "justificativa c": "just_c",
    "justificativa d": "just_d",
    "justificativa e": "just_e",
}


# ── Serializer ────────────────────────────────────────────────────────────────


def _serialize_admin(q: Question) -> dict:
    submitted_tenant = None
    if q.submitted_by_tenant_id:
        t = db.session.get(Tenant, q.submitted_by_tenant_id)
        if t:
            submitted_tenant = {"id": t.id, "name": t.name, "slug": t.slug}

    return {
        "id": q.id,
        "external_id": q.external_id,
        "statement": q.statement,
        "image_url": q.image_url,
        "discipline": q.discipline,
        "topic": q.topic,
        "subtopic": q.subtopic,
        "difficulty": q.difficulty.value if q.difficulty else None,
        "question_type": q.question_type.value if q.question_type else None,
        "correct_alternative_key": q.correct_alternative_key,
        "correct_justification": q.correct_justification,
        "tip": q.tip,
        "exam_board": q.exam_board,
        "exam_year": q.exam_year,
        "exam_name": q.exam_name,
        "review_status": q.review_status.value,
        "rejection_reason": q.rejection_reason,
        "reviewed_at": q.reviewed_at.isoformat() if q.reviewed_at else None,
        "submitted_by_tenant": submitted_tenant,
        "is_active": q.is_active,
        "total_attempts": q.total_attempts,
        "accuracy_rate": q.accuracy_rate,
        "alternatives": [
            {
                "key": a.key,
                "text": a.text,
                "distractor_justification": a.distractor_justification,
            }
            for a in q.alternatives
        ],
        "tags": [t.tag for t in q.tags],
        "created_at": q.created_at.isoformat() if q.created_at else None,
    }


# ── Redis helpers para jobs ───────────────────────────────────────────────────


def _job_key(job_id: str) -> str:
    return f"import_job:{job_id}"


def _job_cancel_key(job_id: str) -> str:
    return f"import_job:{job_id}:cancel"


def _job_get(job_id: str) -> dict | None:
    try:
        raw = redis_client.get(_job_key(job_id))
        return json.loads(raw) if raw else None
    except Exception:
        return None


def _job_set(job_id: str, data: dict):
    try:
        redis_client.setex(_job_key(job_id), _JOB_TTL, json.dumps(data))
    except Exception:
        pass


def _job_is_cancelled(job_id: str) -> bool:
    try:
        return bool(redis_client.exists(_job_cancel_key(job_id)))
    except Exception:
        return False


def _job_set_cancel(job_id: str):
    try:
        redis_client.setex(_job_cancel_key(job_id), _JOB_TTL, "1")
    except Exception:
        pass


# ── Helpers xlsx ──────────────────────────────────────────────────────────────


def _make_external_id(sheet_name: str, row_idx: int, statement: str) -> str:
    raw = f"{sheet_name}::{row_idx}::{statement[:80]}"
    return "xlsx_" + hashlib.md5(raw.encode("utf-8")).hexdigest()[:20]


def _cell_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s.lower() not in ("nan", "none", "") else None


def _get_s3_client():
    region = current_app.config.get("AWS_REGION", "us-east-1")
    return boto3.client(
        "s3",
        region_name=region,
        endpoint_url=f"https://s3.{region}.amazonaws.com",
    )


def _upload_image_to_s3(image_bytes: bytes, filename: str, question_id: str) -> str | None:
    bucket = current_app.config.get("AWS_S3_BUCKET", "")
    region = current_app.config.get("AWS_REGION", "us-east-1")
    if not bucket:
        return None

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "png"
    content_type = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "gif": "image/gif", "webp": "image/webp",
    }.get(ext, "image/png")

    safe_name = filename.replace(" ", "_").lower()
    key = f"questions/images/{question_id}/{safe_name}"

    try:
        s3 = _get_s3_client()
        s3.put_object(Bucket=bucket, Key=key, Body=image_bytes, ContentType=content_type)
        return f"https://{bucket}.s3.{region}.amazonaws.com/{key}"
    except ClientError as e:
        current_app.logger.error(f"S3 image upload error [{filename}]: {e}")
        return None


def _upload_file_to_s3_temp(file_bytes: bytes, job_id: str, ext: str) -> str | None:
    """Salva o arquivo xlsx/zip no S3 temporariamente para o Celery processar."""
    bucket = current_app.config.get("AWS_S3_BUCKET", "")
    region = current_app.config.get("AWS_REGION", "us-east-1")
    if not bucket:
        return None
    key = f"imports/temp/{job_id}.{ext}"
    try:
        s3 = _get_s3_client()
        s3.put_object(Bucket=bucket, Key=key, Body=file_bytes,
                      ContentType="application/octet-stream")
        return key
    except ClientError as e:
        current_app.logger.error(f"S3 temp upload error: {e}")
        return None


def _download_file_from_s3_temp(key: str) -> bytes | None:
    """Baixa o arquivo temporário do S3."""
    bucket = current_app.config.get("AWS_S3_BUCKET", "")
    region = current_app.config.get("AWS_REGION", "us-east-1")
    try:
        s3 = boto3.client("s3", region_name=region,
                          endpoint_url=f"https://s3.{region}.amazonaws.com")
        resp = s3.get_object(Bucket=bucket, Key=key)
        return resp["Body"].read()
    except ClientError:
        return None


def _delete_s3_temp(key: str):
    """Remove o arquivo temporário do S3 após processamento."""
    bucket = current_app.config.get("AWS_S3_BUCKET", "")
    region = current_app.config.get("AWS_REGION", "us-east-1")
    try:
        s3 = boto3.client("s3", region_name=region,
                          endpoint_url=f"https://s3.{region}.amazonaws.com")
        s3.delete_object(Bucket=bucket, Key=key)
    except Exception:
        pass


def _build_embedded_image_map(ws) -> dict[int, tuple[bytes, str]]:
    """
    Escaneia imagens embutidas em uma aba.
    Retorna {excel_row: (bytes, fmt)}.
    anchor._from.row é 0-indexed → excel_row = row + 1.
    """
    image_map: dict[int, tuple[bytes, str]] = {}
    for img in getattr(ws, "_images", []):
        try:
            anch = img.anchor
            if not hasattr(anch, "_from"):
                continue
            excel_row = anch._from.row + 1
            img_bytes = img._data()
            fmt = (getattr(img, "format", None) or "png").lower().strip(".")
            if fmt not in ("png", "jpg", "jpeg", "gif", "webp"):
                fmt = "png"
            image_map[excel_row] = (img_bytes, fmt)
        except Exception:
            continue
    return image_map


def _parse_xlsx_sheets(xlsx_bytes: bytes) -> list[dict]:
    """
    Lê todas as abas do xlsx e retorna lista de dicts normalizados.
    Detecta imagens embutidas automaticamente pela âncora de célula.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        embedded_images = _build_embedded_image_map(ws)

        headers_raw = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_map: dict[int, str] = {}
        for col_idx, raw_header in enumerate(headers_raw):
            normalized = str(raw_header or "").strip().lower()
            if normalized in _XLSX_COL_MAP:
                col_map[col_idx] = _XLSX_COL_MAP[normalized]

        if not col_map:
            continue

        for row_idx in range(2, ws.max_row + 1):
            cells: dict = {}
            for col_idx, field_name in col_map.items():
                cells[field_name] = _cell_str(ws.cell(row=row_idx, column=col_idx + 1).value)

            if not cells.get("statement"):
                continue

            cells["_sheet"] = sheet_name
            cells["_row"] = row_idx

            if row_idx in embedded_images:
                img_bytes, img_fmt = embedded_images[row_idx]
                cells["_embedded_image_bytes"] = img_bytes
                cells["_embedded_image_fmt"] = img_fmt

            rows.append(cells)

    return rows


def _extract_zip(file_bytes: bytes) -> tuple[bytes | None, dict[str, bytes]]:
    xlsx_bytes: bytes | None = None
    image_map: dict[str, bytes] = {}
    image_exts = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        for name in zf.namelist():
            basename = name.rsplit("/", 1)[-1]
            lower = basename.lower()
            if lower.endswith(".xlsx") and not lower.startswith("~$"):
                if xlsx_bytes is None:
                    xlsx_bytes = zf.read(name)
            elif any(lower.endswith(ext) for ext in image_exts):
                image_map[lower] = zf.read(name)

    return xlsx_bytes, image_map


def _insert_or_update_question(
    row: dict,
    image_map: dict[str, bytes],
    enrich_ai: bool,
) -> tuple[str, bool]:
    """
    Insere ou atualiza questão.
    Retorna (status, image_uploaded).
    status: "inserted" | "updated" | "skipped" | "error:mensagem"
    """
    statement = row.get("statement", "")
    sheet_name = row.get("_sheet", "")
    row_idx = int(row.get("_row", 0))

    ext_id = _make_external_id(sheet_name, row_idx, statement)
    q = db.session.query(Question).filter_by(external_id=ext_id).first()

    if q is None:
        if Question.find_duplicate(statement):
            return "skipped", False

    # ── Validações obrigatórias ───────────────────────────────────────────────
    correct_key = (row.get("correct_alternative_key") or "").strip().upper()

    if correct_key not in ("A", "B", "C", "D", "E"):
        return (
            f"error:Gabarito inválido ou ausente: '{correct_key}' "
            f"(esperado A, B, C, D ou E)",
            False,
        )

    alts_present = [k for k in ("a", "b", "c", "d", "e") if row.get(f"alt_{k}")]
    if len(alts_present) < 2:
        return "error:Menos de 2 alternativas preenchidas", False

    if correct_key.lower() not in alts_present:
        return (
            f"error:Gabarito '{correct_key}' não possui alternativa preenchida",
            False,
        )

    is_new = q is None

    if is_new:
        q = Question()
        q.id = str(uuid4())
        q.external_id = ext_id
        q.tenant_id = None
        q.source_type = QuestionSourceType.BANK
        q.review_status = ReviewStatus.APPROVED
        q.submitted_by_tenant_id = None
        q.submitted_by_user_id = None

    q.statement = statement
    q.context = row.get("context")
    q.discipline = (row.get("discipline") or sheet_name or "").strip().upper()
    q.topic = row.get("topic")
    q.subtopic = row.get("subtopic")
    q.tip = row.get("tip")
    q.correct_alternative_key = correct_key
    q.correct_justification = row.get("correct_justification")
    q.is_active = True

    diff_raw = (row.get("difficulty") or "medium").lower().strip()
    q.difficulty = _DIFFICULTY_MAP.get(diff_raw, DifficultyLevel.MEDIUM)

    q.exam_board = row.get("exam_board")
    q.exam_name = row.get("exam_name")
    if row.get("exam_year"):
        try:
            q.exam_year = int(str(row["exam_year"]).split(".")[0])
        except (ValueError, TypeError):
            pass

    # ── Imagem: embutida na célula (prioridade) ou via zip ───────────────────
    image_uploaded = False
    if row.get("_embedded_image_bytes"):
        fmt = row.get("_embedded_image_fmt", "png")
        filename = f"questao_{q.id[:8]}.{fmt}"
        url = _upload_image_to_s3(row["_embedded_image_bytes"], filename, q.id)
        if url:
            q.image_url = url
            image_uploaded = True
    elif row.get("image_file") and image_map:
        img_bytes = image_map.get(row["image_file"].lower())
        if img_bytes:
            url = _upload_image_to_s3(img_bytes, row["image_file"], q.id)
            if url:
                q.image_url = url
                image_uploaded = True

    if is_new:
        db.session.add(q)
        db.session.flush()

        for key in ("a", "b", "c", "d", "e"):
            text = row.get(f"alt_{key}")
            if not text:
                continue
            is_correct = key.upper() == q.correct_alternative_key
            justification = row.get(f"just_{key}") if not is_correct else None
            db.session.add(
                Alternative(
                    id=str(uuid4()),
                    tenant_id=None,
                    question_id=q.id,
                    key=key.upper(),
                    text=text,
                    distractor_justification=justification,
                )
            )

    if enrich_ai or not q.topic or not q.tip:
        try:
            from app.tasks import analyze_question_task
            analyze_question_task.delay(q.id, None)
        except Exception:
            pass

    return ("inserted" if is_new else "updated"), image_uploaded


# ══════════════════════════════════════════════════════════════════════════════
# XLSX PREVIEW
# ══════════════════════════════════════════════════════════════════════════════


@admin_questions_bp.route("/questions/xlsx-preview", methods=["POST"])
@jwt_required()
def xlsx_preview():
    """Analisa o arquivo e retorna resumo sem importar nada."""
    err = _require_super_admin()
    if err:
        return err

    if "file" not in request.files:
        return jsonify(error="Campo 'file' ausente."), 400

    uploaded = request.files["file"]
    filename = (uploaded.filename or "").lower()

    if not filename.endswith((".xlsx", ".zip")):
        return jsonify(error="Formato inválido. Envie um .xlsx ou .zip."), 400

    file_bytes = uploaded.read()
    image_map: dict[str, bytes] = {}

    if filename.endswith(".zip"):
        try:
            xlsx_bytes, image_map = _extract_zip(file_bytes)
        except zipfile.BadZipFile:
            return jsonify(error="Arquivo zip corrompido."), 400
        if not xlsx_bytes:
            return jsonify(error="Nenhum xlsx encontrado no zip."), 400
    else:
        xlsx_bytes = file_bytes

    try:
        rows = _parse_xlsx_sheets(xlsx_bytes)
    except Exception as e:
        return jsonify(error=f"Erro ao ler xlsx: {str(e)}"), 400

    if not rows:
        return jsonify(error="Nenhuma questão encontrada no arquivo."), 400

    by_sheet: dict[str, int] = {}
    disciplines: set[str] = set()
    questions_with_image = 0
    questions_invalid = 0

    for row in rows:
        sheet = row.get("_sheet", "?")
        by_sheet[sheet] = by_sheet.get(sheet, 0) + 1

        has_img = bool(row.get("_embedded_image_bytes")) or bool(row.get("image_file"))
        if has_img:
            questions_with_image += 1

        disc = (row.get("discipline") or sheet).strip().upper()
        if disc:
            disciplines.add(disc)

        correct_key = (row.get("correct_alternative_key") or "").strip().upper()
        alts_present = [k for k in ("a", "b", "c", "d", "e") if row.get(f"alt_{k}")]
        if (
            correct_key not in ("A", "B", "C", "D", "E")
            or len(alts_present) < 2
            or correct_key.lower() not in alts_present
        ):
            questions_invalid += 1

    sample_rows = rows[:100]
    estimated_duplicates = sum(
        1 for r in sample_rows if Question.find_duplicate(r.get("statement", ""))
    )
    if len(rows) > 100:
        estimated_duplicates = int(estimated_duplicates * len(rows) / 100)

    return jsonify(
        total_questions=len(rows),
        by_sheet=[{"sheet": k, "count": v} for k, v in by_sheet.items()],
        disciplines=sorted(disciplines),
        questions_with_image=questions_with_image,
        images_in_zip=len(image_map),
        estimated_duplicates=estimated_duplicates,
        questions_invalid=questions_invalid,
    ), 200


# ══════════════════════════════════════════════════════════════════════════════
# XLSX IMPORT — inicia job assíncrono
# ══════════════════════════════════════════════════════════════════════════════


@admin_questions_bp.route("/questions/xlsx-import", methods=["POST"])
@jwt_required()
def xlsx_import():
    """
    Inicia importação assíncrona via Celery.

    1. Salva o arquivo no S3 (temp)
    2. Cria job no Redis com status "queued"
    3. Dispara Celery task
    4. Retorna job_id imediatamente

    Use GET /questions/import-jobs/<job_id> para acompanhar.
    """
    err = _require_super_admin()
    if err:
        return err

    if "file" not in request.files:
        return jsonify(error="Campo 'file' ausente no form-data."), 400

    uploaded = request.files["file"]
    filename = (uploaded.filename or "").lower()
    enrich_ai = request.form.get("enrich_ai", "true").lower() == "true"

    if not filename.endswith((".xlsx", ".zip")):
        return jsonify(error="Formato inválido. Envie um .xlsx ou .zip."), 400

    file_bytes = uploaded.read()
    ext = "zip" if filename.endswith(".zip") else "xlsx"

    # Valida o arquivo antes de enfileirar
    if ext == "zip":
        try:
            xlsx_bytes, _ = _extract_zip(file_bytes)
        except zipfile.BadZipFile:
            return jsonify(error="Arquivo zip corrompido."), 400
        if not xlsx_bytes:
            return jsonify(error="Nenhum xlsx encontrado no zip."), 400
    else:
        xlsx_bytes = file_bytes

    try:
        rows = _parse_xlsx_sheets(xlsx_bytes)
    except Exception as e:
        return jsonify(error=f"Erro ao ler o xlsx: {str(e)}"), 400

    if not rows:
        return jsonify(error="Nenhuma questão encontrada no arquivo."), 400

    # Salva arquivo no S3 para o Celery processar
    job_id = str(uuid4())
    s3_key = _upload_file_to_s3_temp(file_bytes, job_id, ext)
    if not s3_key:
        return jsonify(error="Erro ao salvar arquivo para processamento."), 500

    # Cria estado inicial do job no Redis
    job_data = {
        "job_id": job_id,
        "status": "queued",          # queued | running | done | cancelled | error
        "filename": uploaded.filename,
        "total": len(rows),
        "processed": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "images_uploaded": 0,
        "enrich_ai": enrich_ai,
        "s3_key": s3_key,
        "s3_ext": ext,
        "error_details": [],
        "started_at": datetime.utcnow().isoformat(),
        "finished_at": None,
    }
    _job_set(job_id, job_data)

    # Dispara task Celery
    try:
        from app.tasks import process_xlsx_import_job
        process_xlsx_import_job.delay(job_id)
    except Exception as e:
        current_app.logger.error(f"Celery dispatch error: {e}")
        return jsonify(error="Erro ao enfileirar processamento."), 500

    return jsonify(job_id=job_id, total=len(rows), status="queued"), 202


# ══════════════════════════════════════════════════════════════════════════════
# JOB STATUS E CONTROLE
# ══════════════════════════════════════════════════════════════════════════════


@admin_questions_bp.route("/questions/import-jobs/<job_id>", methods=["GET"])
@jwt_required()
def get_import_job(job_id: str):
    """Retorna o estado atual de um job de importação."""
    err = _require_super_admin()
    if err:
        return err

    job = _job_get(job_id)
    if not job:
        return jsonify(error="Job não encontrado ou expirado."), 404

    # Calcula percentual
    total = job.get("total", 0)
    processed = job.get("processed", 0)
    job["progress_pct"] = round((processed / total * 100), 1) if total else 0

    return jsonify(job), 200


@admin_questions_bp.route("/questions/import-jobs/<job_id>/cancel", methods=["POST"])
@jwt_required()
def cancel_import_job(job_id: str):
    """Solicita cancelamento de um job em execução."""
    err = _require_super_admin()
    if err:
        return err

    job = _job_get(job_id)
    if not job:
        return jsonify(error="Job não encontrado ou expirado."), 404

    if job.get("status") not in ("queued", "running"):
        return jsonify(error=f"Job já finalizado com status '{job.get('status')}'."), 400

    # Seta flag de cancelamento — o Celery checa a cada 10 linhas
    _job_set_cancel(job_id)

    job["status"] = "cancelling"
    _job_set(job_id, job)

    return jsonify(message="Cancelamento solicitado.", job_id=job_id), 200


@admin_questions_bp.route("/questions/import-jobs", methods=["GET"])
@jwt_required()
def list_import_jobs():
    """
    Lista os últimos jobs de importação (busca por padrão de chave no Redis).
    Retorna até 20 jobs ordenados do mais recente.
    """
    err = _require_super_admin()
    if err:
        return err

    try:
        keys = redis_client.keys("import_job:*")
        # Filtra apenas as chaves de job (não as de cancelamento)
        job_keys = [k for k in keys if not k.endswith(b":cancel")]
    except Exception:
        return jsonify(jobs=[]), 200

    jobs = []
    for key in job_keys:
        try:
            raw = redis_client.get(key)
            if raw:
                job = json.loads(raw)
                total = job.get("total", 0)
                processed = job.get("processed", 0)
                job["progress_pct"] = round((processed / total * 100), 1) if total else 0
                jobs.append(job)
        except Exception:
            continue

    # Ordena por data de início, mais recente primeiro
    jobs.sort(key=lambda j: j.get("started_at", ""), reverse=True)

    return jsonify(jobs=jobs[:20]), 200


# ══════════════════════════════════════════════════════════════════════════════
# BULK IMPORT (JSON — mantido para compatibilidade)
# ══════════════════════════════════════════════════════════════════════════════


@admin_questions_bp.route("/questions/bulk-import", methods=["POST"])
@jwt_required()
def bulk_import():
    err = _require_super_admin()
    if err:
        return err

    data = request.get_json()
    if not isinstance(data, list):
        return jsonify(error="Esperado um array JSON"), 400

    inserted = updated = skipped = errors = 0
    skip_details = []
    error_details = []

    for item in data:
        ext_id = item.get("external_id")
        try:
            q = (
                db.session.query(Question).filter_by(external_id=ext_id).first()
                if ext_id
                else None
            )

            if q is None:
                duplicate = Question.find_duplicate(item.get("statement", ""))
                if duplicate:
                    skip_details.append(
                        {
                            "external_id": ext_id,
                            "reason": "hash_duplicate",
                            "existing_id": duplicate.id,
                        }
                    )
                    skipped += 1
                    continue

            is_new = q is None

            if is_new:
                q = Question()
                q.id = str(uuid4())
                q.external_id = ext_id
                q.tenant_id = None
                q.source_type = QuestionSourceType.BANK
                q.review_status = ReviewStatus.APPROVED
                q.submitted_by_tenant_id = None
                q.submitted_by_user_id = None

            q.statement = item["statement"]
            q.discipline = item.get("discipline", "").strip().upper()
            q.topic = item.get("topic")
            q.subtopic = item.get("subtopic")
            q.difficulty = _DIFFICULTY_MAP.get(
                item.get("difficulty", "medium"), DifficultyLevel.MEDIUM
            )
            q.question_type = _QTYPE_MAP.get(item.get("question_type", ""))
            q.correct_alternative_key = item.get("correct_answer_key", "").upper()
            q.correct_justification = item.get("explanation")
            q.tip = item.get("tip")
            q.exam_board = item.get("exam_board")
            q.exam_year = item.get("exam_year")
            q.exam_name = item.get("exam_name")
            q.source_document_id = item.get("source")
            q.is_active = True

            if is_new:
                db.session.add(q)
                db.session.flush()

                for alt in item.get("alternatives", []):
                    db.session.add(
                        Alternative(
                            id=str(uuid4()),
                            tenant_id=None,
                            question_id=q.id,
                            key=alt["key"].upper(),
                            text=alt["text"],
                            distractor_justification=(
                                alt.get("explanation") if not alt.get("is_correct") else None
                            ),
                        )
                    )

                for tag_str in item.get("tags", []):
                    tag_str = tag_str.strip().lower()
                    if tag_str:
                        db.session.add(
                            QuestionTag(
                                id=str(uuid4()),
                                tenant_id=None,
                                question_id=q.id,
                                tag=tag_str,
                            )
                        )

                inserted += 1
            else:
                updated += 1

            if (inserted + updated) % 50 == 0:
                db.session.commit()

        except Exception as e:
            db.session.rollback()
            errors += 1
            error_details.append({"external_id": ext_id, "error": str(e)})

    db.session.commit()

    return (
        jsonify(
            inserted=inserted,
            updated=updated,
            skipped=skipped,
            errors=errors,
            skip_details=skip_details[:20],
            error_details=error_details[:20],
        ),
        200,
    )


# ══════════════════════════════════════════════════════════════════════════════
# LISTAGEM, REVISÃO E STATS (inalterados)
# ══════════════════════════════════════════════════════════════════════════════


@admin_questions_bp.route("/questions", methods=["GET"])
@jwt_required()
def list_questions():
    err = _require_super_admin()
    if err:
        return err

    discipline = request.args.get("discipline")
    topic = request.args.get("topic")
    difficulty = request.args.get("difficulty")
    q_type = request.args.get("question_type")
    review_status = request.args.get("review_status")
    tenant_filter = request.args.get("submitted_by_tenant_id")
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 30)), 100)

    query = db.session.query(Question).filter(
        Question.source_type == QuestionSourceType.BANK,
        Question.tenant_id.is_(None),
    )
    if discipline:
        query = query.filter(Question.discipline.ilike(f"%{discipline}%"))
    if topic:
        query = query.filter(Question.topic.ilike(f"%{topic}%"))
    if difficulty:
        query = query.filter(Question.difficulty == difficulty)
    if q_type:
        query = query.filter(Question.question_type == q_type)
    if review_status:
        query = query.filter(Question.review_status == review_status)
    if tenant_filter:
        query = query.filter(Question.submitted_by_tenant_id == tenant_filter)

    paginated = query.order_by(Question.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify(
        questions=[_serialize_admin(q) for q in paginated.items],
        total=paginated.total,
        page=page,
        pages=paginated.pages,
    )


@admin_questions_bp.route("/questions/pending", methods=["GET"])
@jwt_required()
def list_pending():
    err = _require_super_admin()
    if err:
        return err

    tenant_filter = request.args.get("tenant_id")
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 30)), 100)

    query = db.session.query(Question).filter(
        Question.review_status == ReviewStatus.PENDING,
        Question.source_type == QuestionSourceType.BANK,
        Question.tenant_id.is_(None),
    )
    if tenant_filter:
        query = query.filter(Question.submitted_by_tenant_id == tenant_filter)

    paginated = query.order_by(Question.created_at.asc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    summary = (
        db.session.query(
            Question.submitted_by_tenant_id,
            func.count(Question.id).label("count"),
        )
        .filter(
            Question.review_status == ReviewStatus.PENDING,
            Question.source_type == QuestionSourceType.BANK,
            Question.tenant_id.is_(None),
            Question.submitted_by_tenant_id.isnot(None),
        )
        .group_by(Question.submitted_by_tenant_id)
        .all()
    )

    by_tenant = []
    for tenant_id, count in summary:
        t = db.session.get(Tenant, tenant_id)
        by_tenant.append(
            {
                "tenant_id": tenant_id,
                "tenant_name": t.name if t else "Desconhecido",
                "tenant_slug": t.slug if t else None,
                "count": count,
            }
        )

    return jsonify(
        questions=[_serialize_admin(q) for q in paginated.items],
        total=paginated.total,
        page=page,
        pages=paginated.pages,
        by_tenant=by_tenant,
    )


@admin_questions_bp.route("/questions/<question_id>/approve", methods=["POST"])
@jwt_required()
def approve_question(question_id):
    err = _require_super_admin()
    if err:
        return err

    q = db.session.get(Question, question_id)
    if not q:
        return jsonify(error="Questão não encontrada"), 404
    if q.source_type != QuestionSourceType.BANK or q.tenant_id is not None:
        return jsonify(error="Apenas questões do banco global podem ser revisadas"), 400

    q.review_status = ReviewStatus.APPROVED
    q.rejection_reason = None
    q.reviewed_by_user_id = get_jwt_identity()
    q.reviewed_at = datetime.utcnow()
    q.is_reviewed = True
    db.session.commit()

    return jsonify(message="Questão aprovada.", question=_serialize_admin(q))


@admin_questions_bp.route("/questions/<question_id>/reject", methods=["POST"])
@jwt_required()
def reject_question(question_id):
    err = _require_super_admin()
    if err:
        return err

    q = db.session.get(Question, question_id)
    if not q:
        return jsonify(error="Questão não encontrada"), 404
    if q.source_type != QuestionSourceType.BANK or q.tenant_id is not None:
        return jsonify(error="Apenas questões do banco global podem ser revisadas"), 400

    data = request.get_json() or {}
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify(error="Motivo da rejeição é obrigatório"), 400

    q.review_status = ReviewStatus.REJECTED
    q.rejection_reason = reason
    q.reviewed_by_user_id = get_jwt_identity()
    q.reviewed_at = datetime.utcnow()
    q.is_reviewed = True
    db.session.commit()

    return jsonify(message="Questão rejeitada.", question=_serialize_admin(q))


@admin_questions_bp.route("/questions/<question_id>", methods=["DELETE"])
@jwt_required()
def delete_question(question_id):
    err = _require_super_admin()
    if err:
        return err

    q = db.session.get(Question, question_id)
    if not q:
        return jsonify(error="Questão não encontrada"), 404

    q.is_active = False
    db.session.commit()
    return jsonify(message="Questão desativada.")


@admin_questions_bp.route("/questions/stats", methods=["GET"])
@jwt_required()
def bank_stats():
    err = _require_super_admin()
    if err:
        return err

    base = db.session.query(Question).filter(
        Question.source_type == QuestionSourceType.BANK,
        Question.tenant_id.is_(None),
    )

    by_status = {
        r[0].value: r[1]
        for r in base.with_entities(Question.review_status, func.count(Question.id))
        .group_by(Question.review_status)
        .all()
    }

    by_discipline = {
        r[0]: r[1]
        for r in base.filter(Question.review_status == ReviewStatus.APPROVED)
        .with_entities(Question.discipline, func.count(Question.id))
        .group_by(Question.discipline)
        .order_by(func.count(Question.id).desc())
        .all()
    }

    by_difficulty = {
        r[0].value: r[1]
        for r in base.filter(Question.review_status == ReviewStatus.APPROVED)
        .with_entities(Question.difficulty, func.count(Question.id))
        .group_by(Question.difficulty)
        .all()
    }

    top_submitters = (
        base.filter(Question.submitted_by_tenant_id.isnot(None))
        .with_entities(
            Question.submitted_by_tenant_id,
            func.count(Question.id).label("total"),
        )
        .group_by(Question.submitted_by_tenant_id)
        .order_by(func.count(Question.id).desc())
        .limit(10)
        .all()
    )

    top_submitters_named = []
    for tenant_id, total in top_submitters:
        t = db.session.get(Tenant, tenant_id)
        top_submitters_named.append(
            {
                "tenant_id": tenant_id,
                "tenant_name": t.name if t else "Desconhecido",
                "total": total,
            }
        )

    return jsonify(
        total_questions=base.count(),
        by_status=by_status,
        by_discipline=by_discipline,
        by_difficulty=by_difficulty,
        top_submitters=top_submitters_named,
    )


# ══════════════════════════════════════════════════════════════════════════════
# REPROCESSAMENTO GEMINI — job assíncrono com progresso em tempo real
# ══════════════════════════════════════════════════════════════════════════════

# Prefixo de chave Redis separado do import job para não misturar históricos
def _reprocess_job_key(job_id: str) -> str:
    return f"reprocess_job:{job_id}"


def _reprocess_cancel_key(job_id: str) -> str:
    return f"reprocess_job:{job_id}:cancel"


def _reprocess_job_get(job_id: str) -> dict | None:
    try:
        raw = redis_client.get(_reprocess_job_key(job_id))
        return json.loads(raw) if raw else None
    except Exception:
        return None


def _reprocess_job_set(job_id: str, data: dict):
    try:
        redis_client.setex(_reprocess_job_key(job_id), _JOB_TTL, json.dumps(data))
    except Exception:
        pass


def _reprocess_job_is_cancelled(job_id: str) -> bool:
    try:
        return bool(redis_client.exists(_reprocess_cancel_key(job_id)))
    except Exception:
        return False


def _reprocess_job_set_cancel(job_id: str):
    try:
        redis_client.setex(_reprocess_cancel_key(job_id), _JOB_TTL, "1")
    except Exception:
        pass


@admin_questions_bp.route("/questions/reprocess-gemini", methods=["POST"])
@jwt_required()
def reprocess_gemini():
    """
    Inicia job assíncrono de reprocessamento Gemini.
    Retorna job_id imediatamente — acompanhe via GET /reprocess-gemini/status.
    """
    err = _require_super_admin()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    limit = int(data.get("limit", 9999))

    # Conta questões pendentes para o job_data
    pending_count = (
        db.session.query(func.count(Question.id))
        .filter(
            Question.source_type == QuestionSourceType.BANK,
            Question.tenant_id.is_(None),
            Question.review_status == ReviewStatus.APPROVED,
            Question.is_active == True,
            Question.tip.is_(None),
        )
        .scalar()
    ) or 0

    already_enriched = (
        db.session.query(func.count(Question.id))
        .filter(
            Question.source_type == QuestionSourceType.BANK,
            Question.tenant_id.is_(None),
            Question.review_status == ReviewStatus.APPROVED,
            Question.tip.isnot(None),
        )
        .scalar()
    ) or 0

    if pending_count == 0:
        return jsonify(
            job_id=None,
            queued=0,
            already_enriched=already_enriched,
            message="Nenhuma questão pendente para reprocessar.",
        ), 200

    job_id = str(uuid4())
    job_data = {
        "job_id": job_id,
        "status": "queued",
        "total": min(pending_count, limit),
        "processed": 0,
        "enriched": 0,
        "skipped": 0,
        "errors": 0,
        "already_enriched": already_enriched,
        "limit": limit,
        "started_at": datetime.utcnow().isoformat(),
        "finished_at": None,
        "error_details": [],
    }
    _reprocess_job_set(job_id, job_data)

    try:
        from app.tasks import run_reprocess_gemini_job
        run_reprocess_gemini_job.delay(job_id, limit)
    except Exception as e:
        current_app.logger.error(f"reprocess_gemini dispatch error: {e}")
        return jsonify(error="Erro ao enfileirar job."), 500

    return jsonify(
        job_id=job_id,
        total=job_data["total"],
        already_enriched=already_enriched,
        status="queued",
    ), 202


@admin_questions_bp.route("/questions/reprocess-gemini/status", methods=["GET"])
@jwt_required()
def reprocess_gemini_status():
    """Retorna o estado atual do job de reprocessamento."""
    err = _require_super_admin()
    if err:
        return err

    job_id = request.args.get("job_id")
    if job_id:
        job = _reprocess_job_get(job_id)
        if not job:
            return jsonify(error="Job não encontrado ou expirado."), 404
        total = job.get("total", 0)
        processed = job.get("processed", 0)
        job["progress_pct"] = round((processed / total * 100), 1) if total else 0
        return jsonify(job), 200

    # Sem job_id → lista histórico (últimos 20)
    try:
        keys = redis_client.keys("reprocess_job:*")
        job_keys = [k for k in keys if not k.endswith(b":cancel")]
    except Exception:
        return jsonify(jobs=[]), 200

    jobs = []
    for key in job_keys:
        try:
            raw = redis_client.get(key)
            if raw:
                job = json.loads(raw)
                total = job.get("total", 0)
                processed = job.get("processed", 0)
                job["progress_pct"] = round((processed / total * 100), 1) if total else 0
                jobs.append(job)
        except Exception:
            continue

    jobs.sort(key=lambda j: j.get("started_at", ""), reverse=True)
    return jsonify(jobs=jobs[:20]), 200


@admin_questions_bp.route("/questions/reprocess-gemini/cancel", methods=["POST"])
@jwt_required()
def reprocess_gemini_cancel():
    """Solicita cancelamento do job de reprocessamento."""
    err = _require_super_admin()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id", "")
    if not job_id:
        return jsonify(error="job_id obrigatório."), 400

    job = _reprocess_job_get(job_id)
    if not job:
        return jsonify(error="Job não encontrado."), 404
    if job.get("status") not in ("queued", "running"):
        return jsonify(error=f"Job já finalizado com status '{job.get('status')}'."), 400

    _reprocess_job_set_cancel(job_id)
    job["status"] = "cancelling"
    _reprocess_job_set(job_id, job)

    return jsonify(message="Cancelamento solicitado.", job_id=job_id), 200