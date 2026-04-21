# api/app/routes/students.py
import re
import secrets
import string
from datetime import datetime, timezone

from flask import Blueprint, jsonify, request
from flask_jwt_extended import get_jwt, jwt_required
from marshmallow import Schema, ValidationError, fields, validate, EXCLUDE

from app.extensions import db, limiter
from app.middleware.tenant import get_current_tenant, require_tenant, resolve_tenant
from app.models.course import Course, CourseEnrollment
from app.models.user import User, UserRole

students_bp = Blueprint("students", __name__)


@students_bp.before_request
def before_request():
    resolve_tenant()


# ── Schemas ───────────────────────────────────────────────────────────────────


class CreateStudentSchema(Schema):
    name = fields.Str(required=True, validate=validate.Length(min=2, max=255))
    email = fields.Email(required=True)
    phone = fields.Str(load_default=None, allow_none=True)
    password = fields.Str(load_default=None, allow_none=True)
    course_ids = fields.List(fields.Str(), load_default=[])

    class Meta:
        unknown = EXCLUDE


class UpdateStudentSchema(Schema):
    name = fields.Str(validate=validate.Length(min=2, max=255))
    is_active = fields.Bool()
    phone = fields.Str(allow_none=True)

    class Meta:
        unknown = EXCLUDE


class BulkStudentItemSchema(Schema):
    name = fields.Str(required=True, validate=validate.Length(min=2, max=255))
    email = fields.Email(required=True)
    phone = fields.Str(load_default=None, allow_none=True)

    class Meta:
        unknown = EXCLUDE


class BulkImportSchema(Schema):
    students = fields.List(fields.Nested(BulkStudentItemSchema), required=True)
    course_ids = fields.List(fields.Str(), load_default=[])

    class Meta:
        unknown = EXCLUDE


# ── Helpers ───────────────────────────────────────────────────────────────────


def _require_producer(claims: dict):
    role = claims.get("role", "")
    if role not in ("producer_admin", "producer_staff", "super_admin"):
        return jsonify({"error": "forbidden", "message": "Acesso restrito a produtores."}), 403
    return None


def _generate_student_password(name: str, phone: str | None) -> str:
    """PrimeiroNome + 4 últimos dígitos do telefone (ou 4 aleatórios)."""
    first_name = name.strip().split()[0].capitalize()
    if phone:
        digits = re.sub(r"\D", "", phone)
        suffix = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
    else:
        suffix = f"{secrets.randbelow(10000):04d}"
    return f"{first_name}{suffix}"


# ── Exportar template de importação ──────────────────────────────────────────


@students_bp.route("/export-template", methods=["GET"])
@jwt_required()
@require_tenant
def export_template():
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    fmt = request.args.get("format", "csv").lower()
    headers = ["Nome", "Email", "Telefone", "Matricular em (nome do curso)"]
    sample = [["Maria Silva", "maria@email.com", "(61) 99999-1234", "Curso de Direito"]]

    if fmt == "json":
        import json as _json
        from flask import Response
        data = [dict(zip(["nome", "email", "telefone", "matricular_em"], row)) for row in sample]
        return Response(
            _json.dumps(data, ensure_ascii=False, indent=2),
            mimetype="application/json",
            headers={"Content-Disposition": "attachment; filename=modelo_alunos.json"},
        )

    if fmt == "xlsx":
        import io
        from flask import Response
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "openpyxl não instalado"}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alunos"
        col_widths = [25, 30, 18, 35]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="4F46E5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]
        ws.row_dimensions[1].height = 22
        for row in sample:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=modelo_alunos.xlsx"},
        )

    # Default: CSV com BOM
    import io as _io
    import csv as _csv
    from flask import Response
    buf = _io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(sample)
    return Response(
        "\ufeff" + buf.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=modelo_alunos.csv"},
    )


# ── Listar alunos ─────────────────────────────────────────────────────────────


@students_bp.route("/", methods=["GET"])
@jwt_required()
@require_tenant
def list_students():
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 20)), 100)
    search = request.args.get("search", "").strip()
    course_id = request.args.get("course_id")

    query = User.query.filter_by(
        tenant_id=tenant.id,
        role=UserRole.STUDENT.value,
        is_deleted=False,
    )

    if search:
        query = query.filter(
            User.name.ilike(f"%{search}%") | User.email.ilike(f"%{search}%")
        )

    if course_id:
        enrolled_ids = [
            e.user_id
            for e in CourseEnrollment.query.filter_by(
                tenant_id=tenant.id,
                course_id=course_id,
                is_active=True,
                is_deleted=False,
            ).all()
        ]
        query = query.filter(User.id.in_(enrolled_ids))

    total = query.count()
    students = query.order_by(User.name).paginate(page=page, per_page=per_page, error_out=False)
    student_ids = [s.id for s in students.items]

    enrollments = CourseEnrollment.query.filter(
        CourseEnrollment.user_id.in_(student_ids),
        CourseEnrollment.tenant_id == tenant.id,
        CourseEnrollment.is_active == True,
        CourseEnrollment.is_deleted == False,
    ).all()

    enrollment_map: dict[str, list[str]] = {}
    for e in enrollments:
        enrollment_map.setdefault(e.user_id, []).append(e.course_id)

    result = []
    for s in students.items:
        prefs = s.preferences or {}
        result.append({
            "id": s.id,
            "name": s.name,
            "email": s.email,
            "phone": prefs.get("phone"),
            "is_active": s.is_active,
            "email_verified": s.email_verified,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "enrolled_course_ids": enrollment_map.get(s.id, []),
        })

    return jsonify({
        "students": result,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": students.pages,
        },
    }), 200


# ── Criar aluno (individual) ──────────────────────────────────────────────────


@students_bp.route("/", methods=["POST"])
@jwt_required()
@require_tenant
@limiter.limit("50 per hour")
def create_student():
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    schema = CreateStudentSchema()
    try:
        data = schema.load(request.get_json(force=True) or {})
    except ValidationError as e:
        return jsonify({"error": "validation_error", "details": e.messages}), 400

    if User.query.filter_by(email=data["email"].lower().strip(), tenant_id=tenant.id, is_deleted=False).first():
        return jsonify({"error": "email_taken", "message": "Já existe um aluno com este e-mail."}), 409

    plain_password = data["password"]
    if not plain_password:
        alphabet = string.ascii_letters + string.digits + "!@#$"
        plain_password = "".join(secrets.choice(alphabet) for _ in range(12))

    preferences = {
        "timezone": "America/Sao_Paulo",
        "notifications_email": True,
        "notifications_push": True,
        "study_reminder_time": "08:00",
    }
    if data.get("phone"):
        preferences["phone"] = data["phone"]

    student = User(
        tenant_id=tenant.id,
        name=data["name"].strip(),
        email=data["email"].lower().strip(),
        role=UserRole.STUDENT,
        is_active=True,
        email_verified=True,
        preferences=preferences,
    )
    student.set_password(plain_password)
    db.session.add(student)
    db.session.flush()

    enrolled_courses = []
    for course_id in data.get("course_ids", []):
        course = Course.query.filter_by(id=course_id, tenant_id=tenant.id, is_deleted=False).first()
        if course:
            db.session.add(CourseEnrollment(
                tenant_id=tenant.id, course_id=course.id, user_id=student.id, is_active=True,
            ))
            enrolled_courses.append(course.name)

    db.session.commit()

    try:
        from app.tasks import send_welcome_email
        import flask
        platform_domain = flask.current_app.config.get("PLATFORM_DOMAIN", "launcheredu.com.br")
        platform_url = f"https://{tenant.slug}.{platform_domain}/login"
        branding = tenant.branding or {}
        send_welcome_email.delay(
            student.email, student.name, plain_password, tenant.name, platform_url,
            course_names=enrolled_courses,
            logo_url=branding.get("logo_url", ""),
            primary_color=branding.get("primary_color", "#4F46E5"),
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"send_welcome_email dispatch falhou: {e}")

    return jsonify({
        "message": "Aluno criado com sucesso.",
        "student": {
            "id": student.id,
            "name": student.name,
            "email": student.email,
            "phone": data.get("phone"),
            "is_active": student.is_active,
            "enrolled_courses": enrolled_courses,
        },
        "temp_password": plain_password if not data["password"] else None,
    }), 201


# ── Importar lista de alunos (bulk JSON) ─────────────────────────────────────


@students_bp.route("/bulk", methods=["POST"])
@jwt_required()
@require_tenant
@limiter.limit("10 per hour")
def bulk_import_students():
    """
    Importa lista de alunos via JSON.
    Processa cada aluno individualmente — erros não bloqueiam os demais.
    """
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    schema = BulkImportSchema()
    try:
        data = schema.load(request.get_json(force=True) or {})
    except ValidationError as e:
        return jsonify({"error": "validation_error", "details": e.messages}), 400

    return _process_bulk(tenant, data["students"], data.get("course_ids", []))


# ── Importar arquivo (CSV / XLSX / JSON) ─────────────────────────────────────


@students_bp.route("/bulk-upload", methods=["POST"])
@jwt_required()
@require_tenant
@limiter.limit("10 per hour")
def bulk_upload_students():
    """
    Importa alunos via upload de arquivo (CSV, XLSX ou JSON).
    O arquivo deve ter colunas: Nome, Email, Telefone (opcional).
    course_ids enviados como form field separado (JSON string).
    """
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()

    if "file" not in request.files:
        return jsonify({"error": "bad_request", "message": "Nenhum arquivo enviado."}), 400

    file = request.files["file"]
    filename = (file.filename or "").lower()

    import json as _json
    course_ids = []
    raw_ids = request.form.get("course_ids", "[]")
    try:
        course_ids = _json.loads(raw_ids)
    except Exception:
        pass

    students_input = []

    try:
        if filename.endswith(".json"):
            import json as _json2
            content = file.read().decode("utf-8")
            rows = _json2.loads(content)
            for row in rows:
                name = row.get("nome") or row.get("name") or row.get("Nome") or ""
                email = row.get("email") or row.get("Email") or ""
                phone = row.get("telefone") or row.get("phone") or row.get("Telefone") or None
                if name and email:
                    students_input.append({"name": name.strip(), "email": email.strip(), "phone": phone})

        elif filename.endswith(".csv"):
            import io as _io
            import csv as _csv
            content = file.read().decode("utf-8-sig")  # remove BOM se presente
            reader = _csv.DictReader(_io.StringIO(content))
            for row in reader:
                name = row.get("Nome") or row.get("nome") or row.get("name") or ""
                email = row.get("Email") or row.get("email") or ""
                phone = row.get("Telefone") or row.get("telefone") or row.get("phone") or None
                if name and email:
                    students_input.append({"name": name.strip(), "email": email.strip(), "phone": phone or None})

        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            import io as _io
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "openpyxl não instalado"}), 500
            wb = openpyxl.load_workbook(_io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                name = str(row_dict.get("Nome") or row_dict.get("nome") or row_dict.get("name") or "").strip()
                email = str(row_dict.get("Email") or row_dict.get("email") or "").strip()
                phone = str(row_dict.get("Telefone") or row_dict.get("telefone") or row_dict.get("phone") or "").strip() or None
                if name and email and email != "None":
                    students_input.append({"name": name, "email": email, "phone": phone})
        else:
            return jsonify({"error": "bad_request", "message": "Formato inválido. Use CSV, XLSX ou JSON."}), 400

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"bulk_upload parse error: {e}")
        return jsonify({"error": "parse_error", "message": f"Erro ao processar arquivo: {str(e)}"}), 400

    if not students_input:
        return jsonify({"error": "bad_request", "message": "Nenhum aluno encontrado no arquivo."}), 400

    if len(students_input) > 500:
        return jsonify({"error": "bad_request", "message": "Máximo de 500 alunos por importação."}), 400

    return _process_bulk(tenant, students_input, course_ids)


# ── Helper de processamento bulk ──────────────────────────────────────────────


def _process_bulk(tenant, students_input: list, course_ids: list):
    """Processa lista de alunos, criando cada um individualmente."""
    import flask

    valid_courses = []
    course_names = []
    for course_id in course_ids:
        course = Course.query.filter_by(id=course_id, tenant_id=tenant.id, is_deleted=False).first()
        if course:
            valid_courses.append(course)
            course_names.append(course.name)

    branding = tenant.branding or {}
    logo_url = branding.get("logo_url", "")
    primary_color = branding.get("primary_color", "#4F46E5")
    platform_domain = flask.current_app.config.get("PLATFORM_DOMAIN", "launcheredu.com.br")
    platform_url = f"https://{tenant.slug}.{platform_domain}/login"

    results = []
    success_count = 0
    error_count = 0

    for idx, item in enumerate(students_input, start=1):
        row_result = {
            "row": idx,
            "name": item.get("name", ""),
            "email": item.get("email", ""),
            "status": None,
            "error": None,
        }

        existing = User.query.filter_by(
            email=item["email"].lower().strip(),
            tenant_id=tenant.id,
            is_deleted=False,
        ).first()

        if existing:
            row_result["status"] = "error"
            row_result["error"] = "E-mail já cadastrado nesta plataforma"
            error_count += 1
            results.append(row_result)
            continue

        plain_password = _generate_student_password(item["name"], item.get("phone"))

        preferences = {
            "timezone": "America/Sao_Paulo",
            "notifications_email": True,
            "notifications_push": True,
            "study_reminder_time": "08:00",
        }
        if item.get("phone"):
            preferences["phone"] = item["phone"]

        try:
            student = User(
                tenant_id=tenant.id,
                name=item["name"].strip(),
                email=item["email"].lower().strip(),
                role=UserRole.STUDENT,
                is_active=True,
                email_verified=True,
                preferences=preferences,
            )
            student.set_password(plain_password)
            db.session.add(student)
            db.session.flush()

            for course in valid_courses:
                db.session.add(CourseEnrollment(
                    tenant_id=tenant.id,
                    course_id=course.id,
                    user_id=student.id,
                    is_active=True,
                ))

            db.session.commit()

            try:
                from app.tasks import send_welcome_email
                send_welcome_email.delay(
                    student.email, student.name, plain_password, tenant.name, platform_url,
                    course_names=course_names, logo_url=logo_url, primary_color=primary_color,
                )
            except Exception:
                pass

            row_result["status"] = "success"
            success_count += 1

        except Exception as e:
            db.session.rollback()
            row_result["status"] = "error"
            row_result["error"] = "Erro interno ao criar aluno"
            error_count += 1
            import logging
            logging.getLogger(__name__).error(f"bulk_import row {idx} error: {e}")

        results.append(row_result)

    return jsonify({
        "results": results,
        "summary": {
            "total": len(students_input),
            "success": success_count,
            "errors": error_count,
            "courses_enrolled": course_names,
        },
    }), 200


# ── Detalhes de um aluno ──────────────────────────────────────────────────────


@students_bp.route("/<string:student_id>", methods=["GET"])
@jwt_required()
@require_tenant
def get_student(student_id: str):
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    student = User.query.filter_by(
        id=student_id, tenant_id=tenant.id, role=UserRole.STUDENT.value, is_deleted=False,
    ).first_or_404()

    enrollments = CourseEnrollment.query.filter_by(
        user_id=student.id, tenant_id=tenant.id, is_active=True, is_deleted=False,
    ).all()

    enrolled_courses = []
    for e in enrollments:
        course = Course.query.filter_by(id=e.course_id, is_deleted=False).first()
        if course:
            enrolled_courses.append({"id": course.id, "name": course.name})

    prefs = student.preferences or {}
    return jsonify({
        "student": {
            "id": student.id,
            "name": student.name,
            "email": student.email,
            "phone": prefs.get("phone"),
            "is_active": student.is_active,
            "email_verified": student.email_verified,
            "created_at": student.created_at.isoformat() if student.created_at else None,
            "enrolled_courses": enrolled_courses,
        }
    }), 200


# ── Editar aluno ──────────────────────────────────────────────────────────────


@students_bp.route("/<string:student_id>", methods=["PUT"])
@jwt_required()
@require_tenant
def update_student(student_id: str):
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    student = User.query.filter_by(
        id=student_id, tenant_id=tenant.id, role=UserRole.STUDENT.value, is_deleted=False,
    ).first_or_404()

    schema = UpdateStudentSchema()
    try:
        data = schema.load(request.get_json(force=True) or {})
    except ValidationError as e:
        return jsonify({"error": "validation_error", "details": e.messages}), 400

    if "name" in data:
        student.name = data["name"].strip()
    if "is_active" in data:
        student.is_active = data["is_active"]
    if "phone" in data:
        prefs = dict(student.preferences or {})
        if data["phone"]:
            prefs["phone"] = data["phone"]
        else:
            prefs.pop("phone", None)
        student.preferences = prefs
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(student, "preferences")

    db.session.commit()
    prefs = student.preferences or {}
    return jsonify({
        "message": "Aluno atualizado.",
        "student": {
            "id": student.id,
            "name": student.name,
            "email": student.email,
            "phone": prefs.get("phone"),
            "is_active": student.is_active,
        },
    }), 200


# ── Gerenciar matrículas de um aluno ──────────────────────────────────────────


@students_bp.route("/<string:student_id>/enrollments", methods=["PUT"])
@jwt_required()
@require_tenant
def manage_student_enrollments(student_id: str):
    claims = get_jwt()
    err = _require_producer(claims)
    if err:
        return err

    tenant = get_current_tenant()
    student = User.query.filter_by(
        id=student_id, tenant_id=tenant.id, role=UserRole.STUDENT.value, is_deleted=False,
    ).first_or_404()

    data = request.get_json(force=True) or {}
    new_course_ids = set(data.get("course_ids", []))

    existing_enrollments = CourseEnrollment.query.filter_by(
        user_id=student.id, tenant_id=tenant.id, is_deleted=False,
    ).all()

    existing_map = {e.course_id: e for e in existing_enrollments}

    for course_id in new_course_ids:
        if course_id in existing_map:
            existing_map[course_id].is_active = True
        else:
            course = Course.query.filter_by(id=course_id, tenant_id=tenant.id, is_deleted=False).first()
            if course:
                db.session.add(CourseEnrollment(
                    tenant_id=tenant.id, course_id=course_id, user_id=student.id, is_active=True,
                ))

    for course_id, enrollment in existing_map.items():
        if course_id not in new_course_ids:
            enrollment.is_active = False

    db.session.commit()
    return jsonify({"message": "Matrículas atualizadas."}), 200